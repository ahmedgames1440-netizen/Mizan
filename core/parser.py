# -*- coding: utf-8 -*-
"""
محرك تحليل نتائج نور — يكتشف صيغة الملف تلقائيًا ويستخرج بيانات الطلاب.
يدعم حاليًا:
  FORMAT_PER_STUDENT_SHEET : كل طالب بصفحة (إشعار درجات نور الرسمي)
  FORMAT_FLAT_TABLE        : جدول مسطّح، عمود واحد لكل طالب (تصدير نور القياسي)
"""
import re
from openpyxl import load_workbook

FORMAT_PER_STUDENT_SHEET = "per_student_sheet"
FORMAT_FLAT_TABLE = "flat_table"
FORMAT_UNKNOWN = "unknown"

NON_SUBJECT_LABELS = {"المواظبة", "السلوك", "مجموع الدرجات الموزونة",
                      "المعدل", "التقدير العام"}


def _extract_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"[\d.]+", str(value))
    return float(m.group()) if m else None


def detect_format(filepath):
    """يفتح الملف ويحدد نوع الصيغة بفحص أول صفحة."""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    # علامة صيغة "صفحة لكل طالب": وجود خلية "اسم الطالب" مصحوبة بعلامات أخرى
    # مميزة لإشعار درجات نور الرسمي (وليس مجرد عمود رأس بصيغة الجدول المسطّح)
    has_name_label = False
    has_per_student_anchor = False
    for row in ws.iter_rows(min_row=1, max_row=35):
        for cell in row:
            if not cell.value:
                continue
            sval = str(cell.value)
            if "اسم الطالب" in sval:
                has_name_label = True
            if sval.strip().startswith("Student's Name") or sval.strip() in (
                "Class :", "Identity No.",
            ):
                has_per_student_anchor = True
    if has_name_label and has_per_student_anchor:
        wb.close()
        return FORMAT_PER_STUDENT_SHEET

    # وإلا، افحص إذا أول صف فيه رؤوس أعمدة متوقعة لصيغة الجدول المسطّح
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [str(c.value).strip() for c in row if c.value]
        joined = " ".join(values)
        if any(k in joined for k in ["الاسم", "الهوية", "الفصل", "المعدل"]):
            header_row = row
            break
    wb.close()
    if header_row is not None:
        return FORMAT_FLAT_TABLE

    return FORMAT_UNKNOWN


def _parse_per_student_sheet(filepath, progress_cb=None):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    sheets = wb.sheetnames
    students = []
    total = len(sheets)

    for idx, sname in enumerate(sheets):
        ws = wb[sname]

        name_ar = None
        name_en = None
        cls = None
        sid = None
        avg = None
        general_grade = None
        rank_grade = None
        rank_class = None
        absence = None
        tardiness = None
        subjects = []

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
        # خريطة: نص التسمية -> (صف، عمود) لإيجاد القيمة المجاورة
        cell_grid = {}
        for row in rows:
            for cell in row:
                if cell.value is not None:
                    cell_grid[(cell.row, cell.column)] = cell.value

        def get(r, c):
            return cell_grid.get((r, c))

        for (r, c), val in list(cell_grid.items()):
            sval = str(val)
            if "اسم الطالب" in sval:
                name_ar = sval.replace("اسم الطالب:", "").strip()
            elif sval.strip().startswith("Student's Name"):
                name_en = sval.replace("Student's Name:", "").strip()
            elif sval.strip() == "Class :":
                for cc in range(c + 1, c + 10):
                    v = get(r, cc)
                    if v is not None:
                        cls = v
                        break
            elif sval.strip() == "Identity No.":
                for cc in range(c + 1, c + 30):
                    v = get(r, cc)
                    if v is not None:
                        sid = v
                        break
            elif sval.strip() == "The Grand Point Average":
                for cc in range(c + 1, c + 10):
                    v = get(r, cc)
                    if v is not None:
                        avg = _extract_number(v)
                        break
            elif sval.strip() == "The General Grade":
                arabic_grades = {"ممتاز", "جيد جداً", "جيد", "مقبول", "ضعيف"}
                for cc in range(c + 1, c + 30):
                    v = get(r, cc)
                    if v is not None and str(v).strip() in arabic_grades:
                        general_grade = v
                        break
            elif sval.strip() == "Sort By Grade :":
                for cc in range(c + 1, c + 5):
                    v = get(r, cc)
                    if v is not None:
                        rank_grade = v
                        break
            elif sval.strip() == "Sort By Class :":
                for cc in range(c + 1, c + 5):
                    v = get(r, cc)
                    if v is not None:
                        rank_class = v
                        break
            elif sval.strip() == "Unexcused absence :":
                for cc in range(c + 1, c + 5):
                    v = get(r, cc)
                    if v is not None:
                        absence = _extract_number(v)
                        break
            elif sval.strip() == "Unexcused tardiness :":
                for cc in range(c + 1, c + 5):
                    v = get(r, cc)
                    if v is not None:
                        tardiness = _extract_number(v)
                        break

        # جدول المواد: نبحث عن صف رأس الجدول "المواد الدراسية" ثم نمسح للأسفل
        header_row_num = None
        col_subject_ar = None
        col_total_pct = None
        col_grade_label = None
        col_final = None
        col_eval_tools = None
        col_short_tests = None

        for (r, c), val in cell_grid.items():
            if str(val).strip() == "المواد الدراسية":
                header_row_num = r
                col_subject_ar = c
            elif str(val).strip() == "التقدير" and header_row_num is None:
                pass

        if header_row_num is not None:
            # ابحث عن أعمدة الرؤوس الأخرى في نفس الصف التقريبي (قد يختلف بصف واحد)
            for (r, c), val in cell_grid.items():
                if abs(r - header_row_num) <= 1:
                    sv = str(val).strip()
                    if sv == "المجموع":
                        col_total_pct = c
                    elif sv == "التقدير":
                        col_grade_label = c
                    elif sv == "اختبار نهاية الفصل":
                        col_final = c
                    elif sv == "أدوات تقييم متنوعة":
                        col_eval_tools = c
                    elif sv == "اختبارات قصيرة":
                        col_short_tests = c

            r = header_row_num + 2
            while True:
                subj_ar = get(r, col_subject_ar)
                if subj_ar is None:
                    break
                subj_ar = str(subj_ar).strip()
                if "مجموع الدرجات الموزونة" in subj_ar:
                    break
                if subj_ar in ("المواظبة", "السلوك"):
                    r += 1
                    if r > header_row_num + 30:
                        break
                    continue
                subjects.append({
                    "subject_ar": subj_ar,
                    "total_pct": _extract_number(get(r, col_total_pct)) if col_total_pct else None,
                    "grade_label": get(r, col_grade_label) if col_grade_label else None,
                    "final_exam": _extract_number(get(r, col_final)) if col_final else None,
                    "eval_tools": _extract_number(get(r, col_eval_tools)) if col_eval_tools else None,
                    "short_tests": _extract_number(get(r, col_short_tests)) if col_short_tests else None,
                })
                r += 1
                if r > header_row_num + 30:
                    break

        if name_ar:
            students.append({
                "sheet": sname,
                "name_ar": name_ar,
                "name_en": name_en,
                "class": cls,
                "id": sid,
                "average": avg,
                "general_grade": general_grade,
                "rank_in_grade": rank_grade,
                "rank_in_class": rank_class,
                "absence": absence,
                "tardiness": tardiness,
                "subjects": subjects,
            })

        if progress_cb:
            progress_cb(idx + 1, total)

    wb.close()
    return students


def _parse_flat_table(filepath, progress_cb=None):
    """صيغة الجدول المسطّح: صف رأس واحد + صف لكل طالب، أعمدة مواد متعددة."""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=False))

    header_row_idx = None
    headers = {}
    for i, row in enumerate(rows[:15]):
        values = {c.column: str(c.value).strip() for c in row if c.value}
        joined = " ".join(values.values())
        if any(k in joined for k in ["الاسم", "اسم الطالب"]) and any(k in joined for k in ["الهوية", "السجل"]):
            header_row_idx = i
            headers = values
            break

    if header_row_idx is None:
        wb.close()
        return []

    col_name = col_id = col_class = col_avg = col_grade = None
    subject_cols = {}
    known_meta = {"الاسم", "اسم الطالب", "الهوية", "رقم الهوية", "السجل المدني",
                  "الفصل", "المعدل", "التقدير", "التقدير العام", "الترتيب",
                  "الغياب", "التأخر", "م"}

    for col, label in headers.items():
        if label in ("الاسم", "اسم الطالب"):
            col_name = col
        elif label in ("الهوية", "رقم الهوية", "السجل المدني"):
            col_id = col
        elif label == "الفصل":
            col_class = col
        elif label == "المعدل":
            col_avg = col
        elif label in ("التقدير", "التقدير العام"):
            col_grade = col
        elif label not in known_meta and label.strip():
            subject_cols[col] = label

    students = []
    data_rows = rows[header_row_idx + 1:]
    total = len(data_rows)
    for idx, row in enumerate(data_rows):
        cellmap = {c.column: c.value for c in row}
        name = cellmap.get(col_name)
        if not name:
            continue
        subjects = []
        for col, label in subject_cols.items():
            v = _extract_number(cellmap.get(col))
            if v is not None:
                subjects.append({"subject_ar": label, "total_pct": v,
                                  "grade_label": None, "final_exam": None,
                                  "eval_tools": None, "short_tests": None})
        students.append({
            "sheet": None,
            "name_ar": str(name).strip(),
            "name_en": None,
            "class": cellmap.get(col_class),
            "id": cellmap.get(col_id),
            "average": _extract_number(cellmap.get(col_avg)),
            "general_grade": cellmap.get(col_grade),
            "rank_in_grade": None,
            "rank_in_class": None,
            "absence": None,
            "tardiness": None,
            "subjects": subjects,
        })
        if progress_cb:
            progress_cb(idx + 1, total)

    wb.close()
    return students


def parse_grades_file(filepath, progress_cb=None):
    """
    نقطة الدخول الرئيسية. تكتشف الصيغة وترجع:
    (قائمة الطلاب, نوع_الصيغة)
    """
    fmt = detect_format(filepath)
    if fmt == FORMAT_PER_STUDENT_SHEET:
        students = _parse_per_student_sheet(filepath, progress_cb)
    elif fmt == FORMAT_FLAT_TABLE:
        students = _parse_flat_table(filepath, progress_cb)
    else:
        students = []
    return students, fmt
